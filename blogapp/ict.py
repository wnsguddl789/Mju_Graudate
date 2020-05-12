from openpyxl import load_workbook
import csv
for year in range (15,21):
    for semester in range(1,3):
        rb = load_workbook("C:/Users/BBAK-JUN/Desktop/information_technology/" + str(year) + '-' + str(semester) + ".xlsx")
        ws = rb['sheet1']

        f = open("./major/ict_"+str(year)+"_"+str(semester)+".csv","w",encoding='UTF-8')
        # csv 형식으로 저장

        f.write("교과목, 학점, 교과목코드\n")
        # 수집한 데이터가 어떤 데이터인 header를 만들어줬음

        subject = []
        point = []
        code =[]
        for i in range(8,150):
            if ws['C'+str(i)].value != None and ws['C'+str(i)].value != "교과목명":
                subject.append(ws['C'+str(i)].value)
                point.append(ws['G'+str(i)].value)
                code.append(ws['D'+str(i)].value)




        for i in range(0,len(subject)-1):
            print("교과목 :",subject[i],"\n학  점 :",point[i],"\n교과목코드 :",code[i],"\n")
            f.write(str(subject[i]) + "," + str(point[i]) + "," + str(code[i]) + "\n")

        f.close()

        if year == 20 and semester == 1: break
